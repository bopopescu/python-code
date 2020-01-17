#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019-10-25 10:27
# @Author  : liujin
# @Email   : 1045833538@QQ.com
# @File    : exc.py
import xlsxwriter
import requests
import openpyxl
import xlrd
from PIL import Image
import os
def geturl():
    # 打开excel文件,获取工作簿对象
    wb = openpyxl.load_workbook('12.xlsx')
    # 从表单中获取单元格的内容
    ws = wb.active  # 当前活跃的表单
    row_range=ws[1:340]
    arr=[]
    for row in range(len(row_range)):  # 打印 2-5行中所有单元格中的值
       arr.append(row_range[row][2].value)

    return arr
def getdata():

    filename = r'12.xlsx'
    datas = xlrd.open_workbook(filename)
    table = datas.sheets()[0]
    arr = []
    for i in range(table.nrows):
        arr.append(table.row_values(i))
    return arr

def getimg(art):
    for i in art:
        if(len(str(i)[31:])>0):
            urlc = i
            pic = requests.get(urlc, timeout=15)
            string = str(i)[31:]+'.jpg'
            import time
            with open("./img/"+string, 'wb') as f:
                f.write(pic.content)
                f.close()
#


def rimg(art,vaule):
    book = xlsxwriter.Workbook('13.xlsx')
    sheet = book.add_worksheet('sheet1')
    sheet.set_column('A:G', 20)


    for t in range(len(vaule)):
        sheet.set_row(t, 200)
        sheet.write_row('A'+str(t), vaule[t])
        if(len(str(art[t])[31:])>0):
            path = os.path.join(os.getcwd(), str("./img/"+str(art[t])[31:]+'.jpg'))
            img = Image.open(path)
            if img.size[0]>1000:
                sheet.insert_image(str('G'+str(t)),str("./img/"+str(art[t])[31:]+'.jpg'),{'x_scale': 0.15, 'y_scale': 0.15})
            if img.size[0]<100:
                    sheet.insert_image(str('G' + str(t)), str("./img/" + str(art[t])[31:] + '.jpg'),{'x_scale': 1, 'y_scale': 1})
            if 100<img.size[0]<1000:
                    sheet.insert_image(str('G' + str(t)), str("./img/" + str(art[t])[31:] + '.jpg'), {'x_scale': 0.25, 'y_scale': 0.25})
    book.close()

# # getimg()
# # rimg()
r = geturl()
# # getimg(r)

rimg(r,getdata())
# print()
